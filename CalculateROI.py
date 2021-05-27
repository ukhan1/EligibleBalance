# -*- coding: utf-8 -*-
"""
Created on Fri Feb 19 15:54:45 2021

@author: Usama
"""

import os
import calendar
import datetime
import sys
import tkinter as tk
from copy import copy
from datetime import date, timedelta
from openpyxl.formula.translate import Translator
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

declared_roi = 1.5
issued = "ap letter issued"
cancelled = "ap letter cancelled"
purchased = "home purchased"

#######################
### Get directories ###
#######################
dir = os.getcwd()
dir_pre = os.path.join(dir, "Before")
dir_post = os.path.join(dir, "After")
dir_error = os.path.join(dir, "error_log.xlsx")
dir_balance = os.path.join(dir, "verify_statements.xlsx")
dir_key = os.path.join(dir,"ACNo_to_File-Mapping.xlsx")
dir_transactions = os.path.join(dir, "Transactions")
dir_transin = os.path.join(dir, "In")
dir_transout = os.path.join(dir, "Out")
if os.path.isfile(dir_error):
    # print ("File exist")
    os.remove(dir_error)
# else:
#     print ("File not exist")
    
if os.path.isfile(dir_balance):
    # print ("File exist")
    os.remove(dir_balance)
# else:
#     print ("File not exist")    
    
error_wb = Workbook()
error_ws = error_wb.active
principal_wb = Workbook()
principal_ws = principal_wb.active
balance_wb = Workbook()
balance_ws = balance_wb.active
dir_tlist = os.listdir(dir_transactions)
dir_list = sorted(os.listdir(dir_pre))
print(dir)
##################################
### Get current quarter months ###
##################################
today = date.today()
tempqtr = ((today.month-1)//3 + 1)
button_exit = 0
def quitButton():
    global button_exit
    button_exit = 1
    master.destroy()
    
def startButton():
    # today = datetime.date(datetime.now())
    global quarter
    global year
    global dir_pre
    global dir_post
    global dir_list
    global dir_transactions
    global dir_transin
    global dir_transout
    global write_principal
    global write_statements
    global compare_statements
    global verify_statements
    global transactions_option
    global declared_roi
    write_principal, write_statements, compare_statements, verify_statements, transactions_option = getBool()
    q,y,r,inp,outp,trns,tinp,toutp = getText()
    try:
        quarter = int(q)
        year = int(y)
        declared_roi = float(r)
        if((quarter <= 0)|(quarter > 4)):
            raise ValueError
        dir_pre = inp
        dir_post = outp
        dir_transactions = trns
        dir_transin = tinp
        dir_transout = toutp
        if(year > today.year):
            print("Invalid year, please try again")
        elif((year == today.year) & (tempqtr <= quarter)):
            print("Invalid quarter, please try again")
        else:
            print("Continuing...")    
            if(os.path.exists(dir_pre)):
                print("continuing...")
                if not os.listdir(dir_pre):
                    print("directory is empty")
                else:
                    print("continuing...")
                    dir_list = sorted(os.listdir(dir_pre))
                    master.destroy()     
            else:
                print("no such path")
    except ValueError:
        print("Invalid quarter or year, please try again")

#Checkbutton Functions
def getBool():
    return var1.get(), var2.get(), var3.get(), var4.get(), var5.get()
def getText():
    global e1
    global e2
    global e3
    global e4
    global e5
    global e6
    global e7
    global e8
    print(var1.get())
    q = e1.get()
    y = e2.get()
    r = e3.get()
    inp= e4.get()
    outp= e5.get()
    trns = e6.get()
    tinp = e7.get()
    toutp = e8.get()
    return q,y,r,inp,outp,trns,tinp,toutp
def click():
    if(var3.get() == False):
        c1.config(state = "normal")
        c2.config(state = "normal")
    else:
        c1.config(state = "disabled")
        c2.config(state = "disabled")
        
#Setting root
master = tk.Tk()
frame = tk.Frame(master)
frame.pack()

#Creating checkbox variables
var1 = tk.BooleanVar(value=True)
var2 = tk.BooleanVar(value=False)
var3 = tk.BooleanVar(value=False)
var4 = tk.BooleanVar(value=False)
var5 = tk.BooleanVar(value=False)
entryr = declared_roi
if (tempqtr == 1):
    entryq = 4
    entryy = today.year-1
else:
    entryq = tempqtr - 1
    entryy = today.year
#Creating and setting frames
middleFrame = tk.Frame(frame)
middleFrame.pack(side = "bottom")
rightFrame1 = tk.Frame(middleFrame)
rightFrame1.pack(side = "right")
rightFrame2 = tk.Frame(rightFrame1)
rightFrame2.pack(side = "right")
rightFrame3 = tk.Frame(rightFrame2)
rightFrame3.pack(side = "bottom")
rightFrame4 = tk.Frame(rightFrame3)
rightFrame4.pack(side = "bottom")
bottomFrame1 = tk.Frame(middleFrame)
bottomFrame1.pack(side = "bottom")
bottomFrame2 = tk.Frame(bottomFrame1)
bottomFrame2.pack(side = "bottom")
bottomFrame3 = tk.Frame(bottomFrame2)
bottomFrame3.pack(side = "bottom") 
bottomFrame4 = tk.Frame(bottomFrame3)
bottomFrame4.pack(side = "bottom") 

#Creating and setting top and middle Labels and Entries
tk.Label(frame, text = "ROI calculator").pack()
tk.Label(middleFrame, text="Quarter").pack(side = "left")
e1 = tk.Entry(middleFrame, width = 1)
e1.insert(0, entryq)
e1.pack(side = "left")
tk.Label(middleFrame, text="Year").pack(side = "left")
e2 = tk.Entry(middleFrame, width = 4)
e2.insert(0, entryy)
e2.pack(side = "left")
tk.Label(middleFrame, text="Roi").pack(side = "left")
e3 = tk.Entry(middleFrame, width = 4)
e3.insert(0, entryr)
e3.pack(side = "left")
tk.Label(middleFrame, text = "%").pack(side = "left")

#Setting 2nd Menu Frames
tk.Label(rightFrame1,text = "").pack(side = "top")
tk.Label(rightFrame1,text = "Transactions").pack(side = "top")
tk.Label(rightFrame1,text = "Input Path").pack(side = "top")
# tk.Label(rightFrame1).pack(side = "top")

tk.Label(rightFrame1, text = "Output Path").pack(side = "top")
c5 = tk.Checkbutton(rightFrame1, text = "Add Transaction", variable = var5)
c5.pack(side = "top")
tk.Label(rightFrame1).pack(side = "top")
tk.Label(rightFrame1).pack(side = "top")
tk.Label(rightFrame1).pack(side = "top")
tk.Label(rightFrame1).pack(side = "top")

tk.Label(rightFrame2, text = "Add Transaction").pack(side = "top")
e6 = tk.Entry(rightFrame2, width = 50)
e6.insert(0, dir_transactions)
e6.pack(side = "top")
e7 = tk.Entry(rightFrame2, width = 50)
e7.insert(0, dir_transin)
e7.pack(side = "top")
e8 = tk.Entry(rightFrame2, width = 50)
e8.insert(0, dir_transout)
e8.pack(side = "top")
# tk.Label(rightFrame2).pack(side = "top")
tk.Label(rightFrame2).pack(side = "top")
tk.Label(rightFrame2).pack(side = "top")
tk.Label(rightFrame2).pack(side = "top")
tk.Label(rightFrame2).pack(side = "top")
tk.Label(rightFrame2).pack(side = "top")
#Creating and setting Input and Output paths
tk.Label(bottomFrame1, text = "Input Path").pack(side = "left")
e4 = tk.Entry(bottomFrame1, width = 50)
e4.insert(0, dir_pre)
e4.pack(side = "right")
tk.Label(bottomFrame2, text = "Output Path").pack(side = "left")
e5 = tk.Entry(bottomFrame2, width = 50)
e5.insert(0, dir_post)
e5.pack(side = "right")


#Creating and setting Checkboxes and Start/Exit buttons
c1 = tk.Checkbutton(bottomFrame3, text = "Write Principal Column", variable = var1)
c1.pack()
c2 = tk.Checkbutton(bottomFrame3, text = "Write Statements", variable = var2)
c2.pack()
c3 = tk.Checkbutton(bottomFrame3, text = "Compare Statements", command = click, variable = var3)
c3.pack()
c4 = tk.Checkbutton(bottomFrame3, text = "Verify Statements", command = click, variable = var4)
c4.pack()
tk.Button(bottomFrame4, text = "Exit", padx = 10, command = quitButton).pack(side = "left")
tk.Button(bottomFrame4, text = "Start", padx = 10, command = startButton).pack(side = "left")

#Quitting after mainloop
tk.mainloop()
master.quit()

if(button_exit == 1):
    print("Exiting")
    sys.exit()
target = str(year) + "Q" + str(quarter)
pstr = target + "principal_column.xlsx"
dir_principal = os.path.join(dir, pstr)
roi_string = "ROI " + target + ": " + "{:.2f}%".format(declared_roi)
print(roi_string)
if (quarter == 1):
    month1 = 1
    month2 = 2
    month3 = 3
elif (quarter == 2):
    month1 = 4
    month2 = 5
    month3 = 6
elif (quarter == 3):
    month1 = 7
    month2 = 8
    month3 = 9
elif (quarter == 4):
    month1 = 10
    month2 = 11
    month3 = 12
else :
    print("Invalid Quarter")
    sys.exit()
    
start_of_m1 = date(year = year, month = month1, day = 1)
start_of_m2 = date(year = year, month = month2, day = 1)
start_of_m3 = date(year = year, month = month3, day = 1)
end_of_m1 =  date(year = year, month = month1, day = calendar.monthrange(year, month1)[1])
end_of_m2 = date(year = year, month = month2, day = calendar.monthrange(year, month2)[1])
end_of_m3 = date(year = year, month = month3, day = calendar.monthrange(year, month3)[1])
end_of_quarter = end_of_m3


print("Month 1:", calendar.month_name[month1], "Last day:", calendar.monthrange(year, month1)[1])
print("Month 2:", calendar.month_name[month2], "Last day:", calendar.monthrange(year, month2)[1])
print("Month 3:", calendar.month_name[month3], "Last day:", calendar.monthrange(year, month3)[1])    
print("End of Quarter:", end_of_quarter)
print(end_of_quarter)

#######################
### File processing ###
#######################

COUNT = 1
def increment():
    global COUNT
    COUNT = COUNT+1
p_COUNT = 1
def p_increment():
    global p_COUNT
    p_COUNT = p_COUNT+1
b_COUNT = 1
def b_increment():
    global b_COUNT
    b_COUNT = b_COUNT+1

def compareStatements(file, calculated, r):
    validation = False;
    matched = False;
    book = load_workbook(os.path.join(dir_post, file))
    ws = book.active
    i = 1
    while(i <= ws.max_row):
        if(ws.cell(row = i, column = 2).value == None):
           i+=1
           continue
        else:
            if(target in ws.cell(row = i, column = 2).value):
                print("Filename:", file, "Roi:", calculated)
                observed = round(ws.cell(row = i, column = 5).value,2)
                principal_ws.cell(row = r, column = 7).value = observed
                print("Checked Value:", observed)
                validation = True
                break
            else:
                i+=1
    if(validation == False):
        print("Values not found. Continuing...")
        return
    if(observed == calculated):
        matched = True;
        principal_ws.cell(row = r, column = 8).value = matched
        print("Values match")
    else:
        matched = False
        principal_ws.cell(row = r, column = 8).value = matched
        print("Values do not match")
    principal_wb.save(dir_principal)
    
def verify_balance(d, file):
    book = load_workbook(os.path.join(d, file), data_only=True)
    ws = book.active
    written_balance = 0
    balance = 0
    r = 9
    while((ws.cell(row = r, column = 7).value != None) | (ws.cell(row = r+1, column = 7).value != None) | (ws.cell(row = r+2, column = 7).value != None)):
        withdrawal = deposit = dividend = roi = 0
        if(ws.cell(row = r, column = 3).value != None):
            if(isinstance(ws.cell(row = r, column = 3).value, float) | isinstance(ws.cell(row = r, column = 3).value, int)):
                withdrawal = ws.cell(row = r, column = 3).value
        if(ws.cell(row = r, column = 4).value != None):
            if(isinstance(ws.cell(row = r, column = 4).value, float) | isinstance(ws.cell(row = r, column = 4).value, int)):
                deposit = ws.cell(row = r, column = 4).value
        if(ws.cell(row = r, column = 5).value != None):
            if(isinstance(ws.cell(row = r, column = 5).value, float) | isinstance(ws.cell(row = r, column = 5).value, int)):
                roi = ws.cell(row = r, column = 5).value
        if(ws.cell(row = r, column = 6).value != None):
            if(isinstance(ws.cell(row = r, column = 6).value, float) | isinstance(ws.cell(row = r, column = 6).value, int)):
                dividend = ws.cell(row = r, column = 6).value
        balance = balance + deposit + roi + dividend - withdrawal
        r+=1
    # if(r > 10):
    written_balance = ws.cell(row = r-1, column = 7).value
   
    written_balance = round(written_balance,2)
    balance = round(balance, 2)
    #     written_balance = ws.cell(row = 9, column = 7).value
    balance_ws.cell(row = b_COUNT, column = 1).value = file
    balance_ws.cell(row = b_COUNT, column = 2).value = balance
    balance_ws.cell(row = b_COUNT, column = 3).value = written_balance
    b_increment()
    balance_wb.save(dir_balance)
    print("Written Balance:", written_balance)
    print("Total Balance:", balance)
    # print("Written Balance:", written_balance)
    return 


### Analyzing entire Transaction file
def add_transaction(file_in):
    book = load_workbook(file_in, data_only = True)
    ws = book.active
    transaction_str = "Deposit_Withdrawal" 
    deposit_str = "Deposit Investor"
    withdraw_str = "Withdrawal Investor"
    #Formatting
    origin = 'G' + str(10)
    formula = "=G9-C10+D10+E10+F10"
    thin = Side(border_style="thin", color="000000")
    # Finding max row 
    z = ws.max_row
    while(ws.cell(row = z, column = 1).value == None):
        z-=1
    # print("Max Row:",z)
    r = 1
    #Until we reach the last transaction
    while(r < z):
        #If it is a withdrawal or deposit, analyze the transaction
        if (transaction_str in ws.cell(row = r, column = 3).value):
            y = 2
            previous_type = 0
            transaction_type = 0
            current_cell = ws.cell(row = r, column = 2).value
            current_date = date(year = current_cell.year, month = current_cell.month, day = current_cell.day)
            account_id = ws.cell(row = r, column = 4).value
            account_name = ws.cell(row = r, column = 5).value
            if(deposit_str in ws.cell(row = r, column = 6).value):
                transaction_type = 1
                description = "Deposit Check"
            elif(withdraw_str in ws.cell(row = r, column = 6).value):
                transaction_type = 2
                description = "Withdrawal Check"
            else:
                print("Transaction Type not correctly specified")
                error_ws.cell(row = COUNT, column = 1).value = file 
                error_ws.cell(row = COUNT, column = 2).value = "Incorrect transaction type"
                error_ws.cell(row = COUNT, column = 3).value = account_name
                error_ws.cell(row = COUNT, column = 4).value = account_id
                error_ws.cell(row = COUNT, column = 5).value = ws.cell(row = r, column = 6).value
                increment()
                error_wb.save(dir_error)
                r+=1
                continue
            # description = ws.cell(row = r, column = 7).value
            document = str(ws.cell(row = r, column = 8).value).strip()
            description = str(description) + str(" # ") + document
            amount = ws.cell(row = r, column = 10).value
            # print("Current date:", current_date)
            print("Account id:", account_id)
            print("Account name:", account_name)
            #Look for the account id
            while((account_ws.cell(row = y, column = 3).value != account_id) & (y <= x)):
                y+=1
            #Match with the filename
            target = account_ws.cell(row = y, column = 1).value
            if(target == None):
                print("No associated account id with that holder")
                error_ws.cell(row = COUNT, column = 1).value = file 
                error_ws.cell(row = COUNT, column = 2).value = "No Matching Account ID"
                error_ws.cell(row = COUNT, column = 3).value = account_name
                error_ws.cell(row = COUNT, column = 4).value = account_id
                increment()
                error_wb.save(dir_error)
            else:
                target_dir = os.path.join(dir_transin, target)
                print(target_dir)
                target_book = load_workbook(target_dir)
                target_ws = target_book.active
                m = target_ws.max_row
                while(target_ws.cell(row = m, column = 1).value == None):
                    m-=1
                #Find the target date
                target_cell = target_ws.cell(row = m, column = 1).value
                target_date = date(year = target_cell.year, month = target_cell.month, day = target_cell.day)
                while((target_date > current_date) & (m!=9)):
                    m-=1
                    target_cell = target_ws.cell(row = m, column = 1).value
                    target_date = date(year = target_cell.year, month = target_cell.month, day = target_cell.day)
                
                #Check for previous deposit/withdrawal type
                if(target_ws.cell(row=m,column=3).value != None):
                    previous_type = 2
                elif(target_ws.cell(row=m,column = 4).value != None):
                    previous_type = 1
                else:
                    #Neither deposit nor withdrawal
                    previous_type = 0
                # print("Target date:", target_date)
                # print("Current date:", current_date)
                #Compare check numbers
                previous_document = str(target_ws.cell(row = m, column = 2).value.partition(" # ")[2]).strip()
                # print("Previous Document:",previous_document)
                # print("Current Document:",document)
                # print("Previous type:", previous_type, "Current Type:",transaction_type)
                if((previous_document == document) & (current_date == target_date)):
                    if(((previous_type == 1) & (transaction_type == 2)) & (document != "AHC")):
                        print("Same day deposit/withdrawal")
                    else:
                        print("Observed duplicate check no.")
                        error_ws.cell(row = COUNT, column = 1).value = file 
                        error_ws.cell(row = COUNT, column = 2).value = "Same date duplicate Check no."
                        error_ws.cell(row = COUNT, column = 3).value = account_name
                        error_ws.cell(row = COUNT, column = 4).value = account_id
                        error_ws.cell(row = COUNT, column = 5).value = target_date
                        error_ws.cell(row = COUNT, column = 6).value = document
                        increment()
                        error_wb.save(dir_error)
                        r+=1
                        continue
                #Insert new row with the extracted information
                target_ws.insert_rows(m+1)
                j = 1
                while(j<9):
                    cell = target_ws.cell(row = 9, column = j)
                    new_cell = target_ws.cell(row=m+1, column = j)
                    if cell.has_style:
                        new_cell._style = copy(cell._style)
                    else:
                        print("No style detected")
                        target_ws.cell(row = m+1, column = j).border = Border(top = thin, right = thin, left = thin, bottom = thin)        
                    new_cell.font = Font(bold = None)
                    j+=1
                
                target_ws.cell(row = m+1, column = 1).value = current_date
                target_ws.cell(row = m+1, column = 2).value = description
                #If its a deposit
                if(transaction_type == 1):
                    target_ws.cell(row = m+1, column = 4).value = amount
                #Else its a withdrawal
                else:
                    target_ws.cell(row = m+1, column = 3).value = amount
                fcell = 'G' + str(m)
                gcell = 'G' + str(m+1)
                target_ws[gcell].value = Translator(formula, origin).translate_formula(fcell)
                if(target_ws[gcell].value != None):
                    target_ws[gcell].value = Translator(formula, origin).translate_formula(gcell)
                target_book.save(os.path.join(dir_transout,target))
                target_book.close()
        # otherwise skip
        r+=1
    print("end of transactions in file")
def partial_roi(ws, file, start_of_m, end_of_m, i, ap):
    #return values: partial roi, error, row, end of file, ap letter issued
    # print("start:", start_of_m, "end:", end_of_m)
    ap_issued = ap
    in_range = 0
    missing_date_error = 0
    unordered_date_error = 0
    cell = ws.cell(row = i, column = 1).value
    minimum = ws.cell(row = i, column = 7).value
    current_date = date(year = cell.year, month = cell.month, day = cell.day)
    if(current_date > start_of_m):
        if(i == 9):
            minimum = 0
            if(end_of_m == end_of_quarter):
                if(ws.cell(row = i+1, column = 1).value is None):
                    return minimum, 0, i, 1, ap_issued
                else:
                    i+=1
            else:
                return minimum, 0, i, 0, ap_issued
    j = i + 1
    while(current_date <= end_of_m):
        try:
            cell = ws.cell(row = i, column = 1).value
            next_cell = ws.cell(row = j, column = 1).value
            current_date = date(year = cell.year, month = cell.month, day = cell.day)
            next_date = date(year = next_cell.year, month = next_cell.month, day = next_cell.day)
            ### comparison error case
            if(next_date < current_date):
                if(unordered_date_error == 0):
                    unordered_date_error = 1
                    print("Unordered dates:", current_date, "and", next_date)
                    if (in_range == 1):
                        error_ws.cell(row = COUNT, column = 1).value = file 
                        error_ws.cell(row = COUNT, column = 2).value = "Dates not in order"
                        increment()
                        error_wb.save(dir_error)
                i=j
                j+=1
                continue
            ### must continue until we reach m
            if(next_date <= start_of_m):
                minimum = ws.cell(row = j, column = 7).value
                if(issued in ws.cell(row = j, column = 2).value.lower()):
                    if(next_date >= (start_of_m - timedelta(days = 90))):
                        print("issued")
                        ap_issued = 1
                elif((cancelled in ws.cell(row = j, column = 2).value.lower()) | (purchased in ws.cell(row = j, column = 2).value.lower())):
                    print("cancelled")
                    ap_issued = 0
                i=j
                j+=1
                continue
            elif((next_date > start_of_m) & (next_date <= end_of_m)):
                in_range = 1
                if(minimum > ws.cell(row = j, column = 7).value):
                    minimum = ws.cell(row = j, column = 7).value
                    i = j
                if(issued in ws.cell(row = j, column = 2).value.lower()):
                    print("ap issued")
                    ap_issued = 1
                    return minimum, 0, i, 0, ap_issued
                elif((cancelled in ws.cell(row = j, column = 2).value.lower()) | (purchased in ws.cell(row = j, column = 2).value.lower())):
                    print("ap_cancelled")
                    ap_issued = 0
                    return 0, 0, i, 0, ap_issued
                if(ws.cell(row = j+1, column = 7).value == None):
                    print("Empty Value")
                    return minimum, 0, j, 0, ap_issued
                j+=1
                continue
            ### outside of range, take last minimum
            elif(next_date > end_of_m):
                # print("end of m", end_of_m, "outside range, j =", j)
                return minimum, 0, j-1, 0, ap_issued
        except AttributeError:
            # print("AttributeError")
            if(ws.cell(row = j+1, column = 7).value != None):
                if(missing_date_error == 0):
                    missing_date_error = 1
                    print("Missing Date")
                    if(in_range == 1):
                        error_ws.cell(row = COUNT, column = 1).value = file 
                        error_ws.cell(row = COUNT, column = 2).value = "Missing Dates"
                        increment()
                        error_wb.save(dir_error)
                        return minimum, 1, i, 0, ap_issued
                j+=1
                continue
            elif(ws.cell(row = j+2, column = 7).value != None):
                if(missing_date_error == 0):
                    missing_date_error = 1
                    print("Missing Date")
                    if(in_range == 1):
                        error_ws.cell(row = COUNT, column = 1).value = file 
                        error_ws.cell(row = COUNT, column = 2).value = "Missing Dates"
                        increment()
                        error_wb.save(dir_error)
                        return minimum, 1, i, 0, ap_issued
                j+=2
                continue
            ### missing dates should be somewhat frequent in the beginning
            else:
                # print("other problem")
                return minimum, 0, i, 1, ap_issued
        except TypeError:
            print("TypeError")
            error_ws.cell(row = COUNT, column = 1).value = file 
            error_ws.cell(row = COUNT, column = 2).value = "Other Error"
            increment()
            error_wb.save(dir_error)
            return 0,1,i,0,ap_issued
        except:
            print("Unexpected error")
            error_ws.cell(row = COUNT, column = 1).value = file 
            error_ws.cell(row = COUNT, column = 2).value = "Unexpected error"
            increment()
            error_wb.save(dir_error)
            return 0,1,i,0,ap_issued
    
    print("Done partial roi")
    # return{'partial_roi' : minimum, 'error' : 0, 'r' : i, 'EoF' : 0}
    return minimum, 0, i, 0, ap_issued

def write_principal_column(file, p1, p2, p3, avg, roi):
    principal_ws.cell(row = p_COUNT, column = 1).value = file
    principal_ws.cell(row = p_COUNT, column = 2).value = p1
    principal_ws.cell(row = p_COUNT, column = 3).value = p2
    principal_ws.cell(row = p_COUNT, column = 4).value = p3
    principal_ws.cell(row = p_COUNT, column = 5).value = avg
    principal_ws.cell(row = p_COUNT, column = 6).value = roi
    principal_wb.save(dir_principal)
def write_output_file(file, file_out, roi, r):
    book = load_workbook(file)
    ws = book.active
    cell = ws.cell(row=r,column=1).value
    current_date = date(year = cell.year, month = cell.month, day = cell.day)
    origin = 'G' + str(10)
    formula = "=G9-C10+D10+E10+F10"
    thin = Side(border_style="thin", color="000000")
    if(current_date > end_of_quarter):
        # print("current date is greater than end of quarter")
        ws.insert_rows(r)
    else:
        # print("current date is not greater than end of quarter, r = ", r+1)
        ws.insert_rows(r+1)
        r+=1
    j = 1
    while(j < 8):
        cell = ws.cell(row = 9, column = j)
        new_cell = ws.cell(row=r, column = j)
        if cell.has_style:
            new_cell._style = copy(cell._style)
        else:
            print("No style detected")
            ws.cell(row = r-1, column = j).border = Border(top = thin, right = thin, left = thin, bottom = thin)        
        new_cell.font = Font(bold = None)
        j +=1
    ws.cell(row = r, column = 1).value = end_of_quarter
    ws.cell(row = r, column = 2).value = roi_string
    ws.cell(row = r, column = 5).value = roi
    ws.cell(row = r, column = 5).alignment = Alignment(horizontal = 'right')
    fcell = 'G' + str(r)
    gcell = 'G' + str(r+1)
    ws[fcell].value = Translator(formula, origin).translate_formula(fcell)
    if(ws[gcell].value != None):
        ws[gcell].value = Translator(formula, origin).translate_formula(gcell)
    print("End of file")
    book.save(file_out)
    book.close()
 
    
def process_file(file_in, file_out, file):
    r = 9
    p1 = 0
    p2 = 0
    p3 = 0
    ap = 0
    os.chdir(dir_pre)
    book = load_workbook(file_in, data_only = True)
    ws = book.active
    z = ws.max_row
    while(ws.cell(row = z, column = 1).value == None):
        z-=1
    print(z)
    cell = ws.cell(row = r, column = 1).value
    try:
        current_date = date(year = cell.year, month = cell.month, day = cell.day)
        if(current_date > end_of_m3):
            print("First date is after the current quarter. Continuing...")
            error_ws.cell(row = COUNT, column = 1).value = file 
            error_ws.cell(row = COUNT, column = 2).value = "After quarter"
            increment()
            error_wb.save(dir_error)
            return
        p1, error, r, EoF, ap = partial_roi(ws, file, start_of_m1, end_of_m1, r, 0)
        p1 = round(p1, 2)
        if(error):
            print("error: check log file")
            return
        if(ap):
            p1 = 0
        if(EoF):
            p2 = p1
            p3 = p1
        else:
            p2, error, r, EoF, ap = partial_roi(ws, file, start_of_m2, end_of_m2, r, ap)
            p2 = round(p2, 2)
            if(error):
                print("error: check log file")
                return
        if(ap):
            p2 = 0
        if(EoF):
            p3 = p2
        else:
            p3, error, r, EoF, ap = partial_roi(ws, file, start_of_m3, end_of_m3, r, ap)
            p3 = round(p3, 2)
            if(error):
                print("error: check log file")
                return
        if(ap):
            p3 = 0     
        print("P1:",p1)
        print("P2:",p2)
        print("P3:",p3)
        avg = (p1 + p2 + p3)/3
        roi = (p1 + p2 + p3) * declared_roi/(3*100)
        roi = round(roi, 2)

    except AttributeError:
       print("Missing date in first row")
       error_ws.cell(row = COUNT, column = 1).value = file 
       error_ws.cell(row = COUNT, column = 2).value = "Missing first row"
       increment()
       error_wb.save(dir_error)
       book.close()
       return
    except TypeError:
       print("Possible incorrect value")
       error_ws.cell(row = COUNT, column = 1).value = file 
       error_ws.cell(row = COUNT, column = 2).value = "Incorrect value"
       increment()
       error_wb.save(dir_error)
       book.close()
       return
    except KeyboardInterrupt:
        print("Exiting...")
        book.close()
        sys.exit()
    except:
       print("Unexpected error")
       error_ws.cell(row = COUNT, column = 1).value = file 
       error_ws.cell(row = COUNT, column = 2).value = "Unexpected error"
       increment()
       error_wb.save(dir_error)
       book.close() 
       return
    if(write_principal == True):
        write_principal_column(file, p1, p2, p3, avg, roi)
        p_increment()
    if(write_statements == True):
        write_output_file(file, file_out, roi, r)
    if(verify_statements == True):
        if(write_statements == True):
            verify_balance(dir_post, file)
        else:
            verify_balance(dir_pre, file)
    book.close()
    
if(write_principal == True):
    principal_ws.cell(row =  1, column = 1).value = "File No"
    principal_ws.cell(row =  1, column = 2).value = "P1"
    principal_ws.cell(row =  1, column = 3).value = "P2"
    principal_ws.cell(row =  1, column = 4).value = "P3"
    principal_ws.cell(row =  1, column = 5).value = "Avg"
    principal_ws.cell(row =  1, column = 6).value = "Calculated ROI"
    principal_ws.cell(row =  1, column = 7).value = "Observed ROI"
    principal_ws.cell(row =  1, column = 8).value = "Values Match?"
    p_increment()
    
if(verify_statements == True):
    balance_ws.cell(row = 1, column = 1).value = "File"
    balance_ws.cell(row = 1, column = 1).value = "Calculated Balance"
    balance_ws.cell(row = 1, column = 1).value = "Written Balance"
    b_increment()

if(transactions_option == True):
    account_book = load_workbook(dir_key)
    account_ws = account_book.active
    x = account_ws.max_row
    while(account_ws.cell(row = x, column = 1).value == None):
        x-=1
    print(x, "amount of accounts")
    print(account_ws.cell(row = 1, column = 1).value)
    for file in filter(lambda x: not (x.startswith('~') or x.startswith('.')), dir_tlist):
        file_in = os.path.join(dir_transactions, file)
        print("Transaction file:", file_in)
        add_transaction(file_in)   
    account_book.close()
if(compare_statements == True):
    principal_wb = load_workbook(dir_principal)
    principal_ws = principal_wb.active
    row_count = principal_ws.max_row
    column_count = principal_ws.max_column
    i = 2
    while(i <= row_count):
        fileno = principal_ws.cell(row = i, column = 1).value
        calculated_roi = principal_ws.cell(row = i, column = 6).value
        compareStatements(fileno, calculated_roi, i)
        i+=1
else:
    if(write_principal | write_statements | verify_statements | compare_statements):
        for file in sorted( filter(lambda x: not (x.startswith('~') or x.startswith('.')), dir_list) ):
            file_in = os.path.join(dir_pre, file)
            file_out = os.path.join(dir_post, file)
            #file_in = file_in.replace(os.sep, '/')
            print(file_in)
            process_file(file_in, file_out, file)   
            
    
        
        
        























